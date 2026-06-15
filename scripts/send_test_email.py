#!/usr/bin/env python3
"""
send_test_email.py
==================
VM 提供测试题附件 → 我分析附件内容生成摘要 → 展示邮件全文 → VM 确认 → 发送

用法：
    python3 scripts/send_test_email.py --list
    python3 scripts/send_test_email.py --name "测试候选人A" --file ~/Downloads/test.pdf
    python3 scripts/send_test_email.py --name "Miroslaw" --file test.pdf --dry-run
    python3 scripts/send_test_email.py --name "测试候选人A" --file test.pdf --yes
"""

import sys, re, json, argparse, subprocess, base64, shutil, time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
from config_loader import load_config, get_smtp, get_lark, get_paths, is_test_mode, get_test_email
from field_resolver import field_id_or
from lark_cli_utils import normalize_record_list_data, run_lark_cli_json
from manual_trace import log_manual_step

_CFG = load_config()

# ── 配置 ──────────────────────────────────────────────────────────────────────
BASE_TOKEN = get_lark(_CFG).get("base_token", "")
TABLE_ID   = get_lark(_CFG).get("resume_table_id", "")

SMTP_HOST  = get_smtp(_CFG).get("host", "")
SMTP_PORT  = get_smtp(_CFG).get("port", 465)
SMTP_USER  = get_smtp(_CFG).get("user", "")
SMTP_PASS  = get_smtp(_CFG).get("password", "")

TEST_MODE  = is_test_mode(_CFG)
TEST_EMAIL = get_test_email(_CFG)

FLD_NAME         = field_id_or("candidate", "candidate.name", "fldSAfsOJf")
FLD_EMAIL        = field_id_or("candidate", "candidate.email", "fldWf5X8NR")
FLD_LANG_PAIR    = field_id_or("candidate", "candidate.language_pair", "fldBvHUo5K")
FLD_STATUS       = field_id_or("candidate", "candidate.status", "fldfp6Pn7l")
FLD_TEST_SENT_AT = field_id_or("candidate", "candidate.test_sent_at", "fldQLxyrP7")

VALID_EXTS = {".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc", ".txt", ".xlsx"}

LANG_HINTS = [
    {
        "tokens": ["中译韩", "中韩", "韩语", "Korean", "ko"],
        "expected_any": ["韩", "Korean", "ko"],
        "label": "中译韩/韩语",
    },
    {
        "tokens": ["中译英", "中英", "英语", "English", "en"],
        "expected_any": ["英", "English", "en"],
        "label": "中译英/英语",
    },
    {
        "tokens": ["中译日", "中日", "日语", "Japanese", "ja"],
        "expected_any": ["日", "Japanese", "ja"],
        "label": "中译日/日语",
    },
]

# ── lark-cli ──────────────────────────────────────────────────────────────────
def lark_cli(*args):
    resp = run_lark_cli_json(*args)
    if not isinstance(resp, dict):
        raise RuntimeError(f"非 JSON 返回:\n{str(resp)[:200]}")
    return resp

def extract_text(val):
    if not val: return ""
    if isinstance(val, list):
        parts = []
        for v in val:
            if isinstance(v, dict):
                parts.append(v.get("name") or v.get("text") or "")
            else:
                parts.append(str(v))
        text = " ".join(p for p in parts if p).strip()
    else:
        text = str(val).strip()
    # 清洗 markdown 链接格式：[text](url) → text
    import re
    text = re.sub(r'\[([^\]]+)\]\([^)]*\)', r'\1', text)
    return text

def fetch_records():
    records, page_token = [], None
    while True:
        args = ["base", "+record-list", "--base-token", BASE_TOKEN,
                "--table-id", TABLE_ID, "--format", "json", "--limit", "100"]
        if page_token:
            args += ["--page-token", page_token]
        resp = lark_cli(*args)
        db = resp["data"]
        records.extend(normalize_record_list_data(db))
        if not db.get("has_more") or not db.get("page_token"):
            break
        page_token = db["page_token"]
    return records

def update_record(record_id, fields):
    resp = lark_cli("base", "+record-upsert", "--base-token", BASE_TOKEN,
                    "--table-id", TABLE_ID, "--record-id", record_id,
                    "--json", json.dumps(fields, ensure_ascii=False))
    if not resp.get("ok"):
        raise RuntimeError(f"写回失败: {resp.get('error')}")

# ── 附件分析 ──────────────────────────────────────────────────────────────────
def summarize_attachment(file_path: Path) -> str:
    """
    分析附件内容，生成摘要供 VM 确认。
    PDF/图片 → 用视觉模型；文本类 → 直接读取前2000字符。
    返回摘要字符串。
    """
    ext = file_path.suffix.lower()

    # 文本类：直接读取
    if ext in (".txt",):
        try:
            content = file_path.read_text(encoding="utf-8", errors="replace")[:2000]
            lines = content.strip().splitlines()
            preview = "\n".join(f"    {l}" for l in lines[:20])
            return f"【文本内容预览（前20行）】\n{preview}"
        except Exception as e:
            return f"【读取失败】{e}"

    # Excel → 解析内容结构
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            lines = []
            total_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                non_empty = [(r, row) for r, row in enumerate(ws.iter_rows(values_only=True), 1)
                             if any(c is not None for c in row)]
                total_rows += len(non_empty)
                lines.append(f"Sheet「{sheet_name}」：{ws.max_row} 行 × {ws.max_column} 列")

                # 识别题目结构：找列标题行
                header_row = None
                sections = {}   # section_name -> count
                current_section = None
                source_col = None

                for r, row in non_empty:
                    row_text = [str(c).strip() for c in row if c is not None]
                    joined = " ".join(row_text)

                    # 识别列标题行（含「原文」「Source」「简中」等）
                    if any(kw in joined for kw in ["原文", "Source", "简中", "SRC", "原文本"]):
                        header_row = r
                        # 找原文列索引
                        for ci, c in enumerate(row):
                            if c and any(kw in str(c) for kw in ["原文", "Source", "简中", "SRC", "原文本"]):
                                source_col = ci
                        continue

                    # 识别段落标题（单列非空、短文本）
                    non_none = [c for c in row if c is not None]
                    if len(non_none) == 1 and len(str(non_none[0])) < 30:
                        txt = str(non_none[0]).strip()
                        if txt and not txt.startswith("·") and not txt.startswith("*") and not txt.startswith("ex"):
                            current_section = txt
                            if current_section not in sections:
                                sections[current_section] = 0
                            continue

                    # 统计内容行
                    if header_row and r > header_row and current_section:
                        if source_col is not None and row[source_col] is not None:
                            sections[current_section] = sections.get(current_section, 0) + 1
                        elif source_col is None and non_none:
                            sections[current_section] = sections.get(current_section, 0) + 1

                # 输出结构摘要
                if sections:
                    lines.append("  题目结构：")
                    for sec, cnt in sections.items():
                        if cnt > 0:
                            lines.append(f"    · {sec}：{cnt} 条")
                        else:
                            lines.append(f"    · {sec}")

                # 显示前3条原文样例
                if header_row and source_col is not None:
                    samples = []
                    for r, row in non_empty:
                        if r > header_row and source_col < len(row) and row[source_col]:
                            txt = str(row[source_col]).strip()
                            if txt and len(txt) > 2:
                                samples.append(txt)
                        if len(samples) >= 3:
                            break
                    if samples:
                        lines.append("  原文样例：")
                        for s in samples:
                            preview = s[:40] + "…" if len(s) > 40 else s
                            lines.append(f"    · {preview}")

            size_kb = file_path.stat().st_size // 1024
            lines.insert(0, f"【Excel 附件】{file_path.name}，{size_kb} KB，共 {total_rows} 条有效行")
            return "\n".join(lines)
        except ImportError:
            size_kb = file_path.stat().st_size // 1024
            return f"【Excel 文件】{file_path.name}，{size_kb} KB（未安装 openpyxl，无法解析内容）"
        except Exception as e:
            return f"【Excel 解析失败】{e}"

    # PDF / 图片 → base64 → 视觉模型
    if ext in (".pdf", ".jpg", ".jpeg", ".png", ".webp"):
        try:
            # PDF 转第一页图片（需 ImageMagick）
            img_path = file_path
            if ext == ".pdf":
                tmp = file_path.with_suffix("._p0.png")
                r = subprocess.run(
                    ["convert", "-density", "120", f"{file_path}[0]", str(tmp)],
                    capture_output=True
                )
                if r.returncode == 0:
                    img_path = tmp
                else:
                    return "【PDF 预览】无法转换图片（未安装 ImageMagick），请 VM 确认附件内容正确后继续"

            # 调用视觉模型（写到临时文件，主 session 的 image tool 读取）
            # 这里返回路径标记，由主 session 处理
            return f"__VISION__{img_path}"
        except Exception as e:
            return f"【分析失败】{e}"

    # Word/其他 → 只说明文件名和大小
    size_kb = file_path.stat().st_size // 1024
    return f"【{ext.upper()} 文件】{file_path.name}，{size_kb} KB\n    （内容需 VM 自行确认正确）"

# ── 邮件构建 ──────────────────────────────────────────────────────────────────
def build_email_content(name: str, lang_pair: str, lang: str, attachment_name: str) -> tuple[str, str]:
    if lang == "zh":
        subject = f"【37GAMES】翻译能力测试 - {name}"
        body = f"""您好，{name}，

感谢您对本次翻译合作的兴趣！

我们已审阅您的简历，希望进一步了解您的翻译水平。请查收附件中的翻译测试题（{attachment_name}）。

测试说明：
- 语言方向：{lang_pair or '请参考附件说明'}
- 请在收到后 5 个工作日内完成并将译文发回此邮箱
- 如有任何疑问，欢迎随时联系

期待您的回复！

此致
37GAMES 本地化团队"""
    else:
        subject = f"[37GAMES] Translation Test - {name}"
        body = f"""Dear {name},

Thank you for your interest in collaborating with us!

We have reviewed your resume and would like to assess your translation skills. Please find the attached translation test ({attachment_name}).

Test details:
- Language pair: {lang_pair or 'Please refer to the attachment'}
- Please complete and reply within 5 business days
- Feel free to reach out if you have any questions

We look forward to hearing from you!

Best regards,
37GAMES Localization Team"""
    return subject, body

# ── 发送邮件 ──────────────────────────────────────────────────────────────────
def build_email_message(to_email: str, subject: str, body: str, attachment: Path):
    """构建 MIMEMultipart 邮件对象（不发送）"""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    actual_to = TEST_EMAIL if TEST_MODE else to_email
    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = actual_to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(attachment, "rb") as fh:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{attachment.name}"')
    msg.attach(part)
    return msg, actual_to


def save_draft(msg, draft_dir: Path, filename: str):
    """保存为 .eml 草稿文件，VM 双击用邮件客户端打开后点发送"""
    draft_dir.mkdir(parents=True, exist_ok=True)
    draft_path = draft_dir / filename
    draft_path.write_text(msg.as_string(), encoding="utf-8")
    return draft_path


def send_email(to_email: str, subject: str, body: str, attachment: Path, draft: bool = False):
    import smtplib, ssl

    msg, actual_to = build_email_message(to_email, subject, body, attachment)

    if draft:
        draft_dir = Path(get_paths(_CFG).get("contract_output", "~/Documents/loc-contracts/output/")).expanduser() / "drafts"
        safe_name = re.sub(r'[^\w\-\u4e00-\u9fff.]', '_', to_email)
        draft_path = save_draft(msg, draft_dir, f"测试题_{safe_name}.eml")
        print(f"📝 草稿已保存：{draft_path}")
        print(f"   双击用邮件客户端打开，确认无误后点发送")
        return

    ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    ctx.check_hostname = False
    ctx.verify_mode    = ssl.CERT_NONE
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as srv:
        srv.login(SMTP_USER, SMTP_PASS)
        srv.sendmail(SMTP_USER, actual_to, msg.as_string())

    tag = f"⚠️  [测试模式] 发到 {actual_to}（原始目标：{to_email}）" if TEST_MODE else f"✅ 已发至 {actual_to}"
    print(tag)

# ── 列表 ─────────────────────────────────────────────────────────────────────
def list_records(records):
    print(f"{'#':<4} {'record_id':<22} {'姓名':<20} {'邮箱':<32} {'招募状态'}")
    print("-" * 100)
    for i, rec in enumerate(records, 1):
        f = rec["fields"]
        print(f"{i:<4} {rec['record_id']:<22} "
              f"{extract_text(f.get(FLD_NAME)):<20} "
              f"{extract_text(f.get(FLD_EMAIL)):<32} "
              f"{extract_text(f.get(FLD_STATUS))}")


def detect_attachment_language_warning(file_path: Path, attachment_summary: str, candidate_lang_pair: str) -> str:
    """Return a warning when the attachment language signal conflicts with the candidate language pair."""
    haystack = f"{file_path.name}\n{attachment_summary}"
    for hint in LANG_HINTS:
        if any(token.lower() in haystack.lower() for token in hint["tokens"]):
            if not any(token.lower() in candidate_lang_pair.lower() for token in hint["expected_any"]):
                return (
                    f"附件疑似为「{hint['label']}」测试题，但候选人语言对是「{candidate_lang_pair}」。"
                    "请确认是否选错测试题。"
                )
    return ""

# ── 主流程 ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="测试题邮件发送")
    parser.add_argument("--name",      help="资源商姓名（模糊匹配）")
    parser.add_argument("--record-id", help="飞书 record_id（精确）")
    parser.add_argument("--file",      metavar="PATH", help="测试题附件路径")
    parser.add_argument("--list",      action="store_true")
    parser.add_argument("--dry-run",   action="store_true", help="预览但不发送")
    parser.add_argument("--prepare",   action="store_true", help="仅输出可复制邮件包，不发送、不写状态")
    parser.add_argument("--yes",       action="store_true", help="跳过确认直接发送")
    parser.add_argument("--draft",     action="store_true", help="保存草稿而非直接发送，VM 双击 .eml 后点发送")
    parser.add_argument("--allow-language-mismatch", action="store_true", help="允许附件语言方向与候选人语言对不一致")
    args = parser.parse_args()

    records = fetch_records()

    if args.list:
        list_records(records)
        return

    if not args.name and not args.record_id:
        parser.print_help(); sys.exit(0)

    if not args.file:
        print("❌ 请用 --file 指定测试题附件路径")
        sys.exit(1)

    file_path = Path(args.file).expanduser().resolve()
    if not file_path.exists():
        print(f"❌ 文件不存在：{file_path}")
        sys.exit(1)
    if file_path.stat().st_size == 0:
        print(f"❌ 文件为空：{file_path}")
        print("   请重新下载或确认测试题附件后再发送，避免给候选人发送空文件。")
        sys.exit(1)
    if file_path.suffix.lower() not in VALID_EXTS:
        print(f"❌ 不支持的格式：{file_path.suffix}（支持：{' / '.join(VALID_EXTS)}）")
        sys.exit(1)

    # 找候选人
    target = None
    if args.record_id:
        for r in records:
            if r["record_id"] == args.record_id:
                target = r; break
    elif args.name:
        matches = [r for r in records
                   if args.name.lower() in extract_text(r["fields"].get(FLD_NAME)).lower()]
        if not matches:
            print(f"❌ 未找到：{args.name}"); sys.exit(1)
        if len(matches) > 1:
            print(f"⚠️  找到 {len(matches)} 条，请用 --record-id 精确指定：")
            list_records(matches); sys.exit(1)
        target = matches[0]

    if not target:
        print(f"❌ 未找到记录"); sys.exit(1)

    f         = target["fields"]
    name      = extract_text(f.get(FLD_NAME)) or "未知"
    email     = extract_text(f.get(FLD_EMAIL))
    lang_pair = extract_text(f.get(FLD_LANG_PAIR))
    lang      = "zh" if re.search(r'[\u4e00-\u9fff]', name) else "en"

    print(f"\n候选人：{name}  ({email})  语言对：{lang_pair}\n")

    # ── 附件摘要 ──────────────────────────────────────────────────────────────
    print("分析附件内容...")
    summary = summarize_attachment(file_path)
    if "解析失败" in summary or "读取失败" in summary:
        print(f"❌ 附件无法解析：{summary}")
        print("   请重新下载或确认测试题附件后再发送，避免给候选人发送损坏文件。")
        sys.exit(1)

    # 视觉模型处理（返回路径标记时由主 session image tool 完成）
    attachment_summary = summary
    if summary.startswith("__VISION__"):
        img_path = summary.replace("__VISION__", "")
        # 输出标记供主 session 调用 image tool
        print(f"__NEEDS_VISION__:{img_path}")
        attachment_summary = f"【附件预览待视觉分析】路径：{img_path}"

    lang_warning = detect_attachment_language_warning(file_path, attachment_summary, lang_pair)

    # ── 构建邮件 ──────────────────────────────────────────────────────────────
    subject, body = build_email_content(name, lang_pair, lang, file_path.name)

    # ── 展示完整摘要 ──────────────────────────────────────────────────────────
    print("=" * 62)
    print("📎 附件内容摘要")
    print("=" * 62)
    print(attachment_summary)
    print()
    print("=" * 62)
    print("📧 邮件预览")
    print("=" * 62)
    print(f"收件人：{TEST_EMAIL if TEST_MODE else email}")
    print(f"附  件：{file_path.name}  ({file_path.stat().st_size // 1024} KB)")
    print(f"主  题：{subject}")
    print("-" * 62)
    print(body)
    print("=" * 62)

    if lang_warning:
        print(f"\n⚠️  语言方向风险：{lang_warning}")
        if not args.dry_run and not args.allow_language_mismatch:
            print("❌ 已阻断发送。若确认要发送，请显式加 --allow-language-mismatch。")
            sys.exit(1)

    if TEST_MODE:
        print(f"\n⚠️  [测试模式] 实际发到：{TEST_EMAIL}（而非 {email}）")

    if args.dry_run:
        print("\n[DRY-RUN] 不发送")
        log_manual_step(
            step_name="测试题邮件 dry-run",
            status="skipped",
            candidate_name=name,
            candidate_record_id=target["record_id"],
            input_summary=f"附件: {file_path.name}; 语言对: {lang_pair}",
            output_summary=f"收件人(TEST_MODE): {TEST_EMAIL if TEST_MODE else email}",
        )
        return

    if args.prepare:
        print("\n[PREPARE] 已生成邮件包，不发送、不写入飞书状态")
        log_manual_step(
            step_name="测试题邮件准备",
            status="skipped",
            candidate_name=name,
            candidate_record_id=target["record_id"],
            input_summary=f"附件: {file_path.name}; 语言对: {lang_pair}",
            output_summary="已输出邮件标题、正文、附件路径；未发送，未写状态",
        )
        return

    if not args.yes:
        print()
        ans = input("确认发送以上邮件（含附件）？[y/N] ").strip().lower()
        if ans != "y":
            print("❌ 已取消"); sys.exit(0)

    print("\n发送中...")
    send_email(email, subject, body, file_path, draft=args.draft)

    if args.draft:
        print("📝 已保存本地草稿，不更新飞书状态；VM 发送后请再推进状态。")
        log_manual_step(
            step_name="测试题邮件草稿",
            status="skipped",
            candidate_name=name,
            candidate_record_id=target["record_id"],
            input_summary=f"附件: {file_path.name}; 语言对: {lang_pair}",
            output_summary="本地 .eml 草稿已保存；未发送，未写状态",
        )
        return

    print("更新飞书状态...")
    update_record(target["record_id"], {
        FLD_STATUS:       "📤 测试中",
        FLD_TEST_SENT_AT: int(time.time() * 1000),
    })
    print("✅ 招募状态 → 📤 测试中，测试发送时间已记录")
    log_manual_step(
        step_name="测试题邮件发送",
        status="done",
        candidate_name=name,
        candidate_record_id=target["record_id"],
        input_summary=f"附件: {file_path.name}; 语言对: {lang_pair}",
        output_summary=f"发送至: {TEST_EMAIL if TEST_MODE else email}; 状态=📤 测试中",
    )

if __name__ == "__main__":
    main()
